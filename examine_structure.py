#!/usr/bin/env python3
"""
Examine the structure of both Excel files to understand the layout
"""

import openpyxl
from pathlib import Path

def examine_structure(file_path, file_name):
    """Examine the structure of the Excel file"""
    print(f"\n📊 EXAMINING {file_name}")
    print("=" * 60)
    
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
        
        # Show headers (first row)
        print(f"\nHeaders (Row 1):")
        headers = []
        for col in range(1, min(sheet.max_column + 1, 24)):  # Show first 23 columns
            cell = sheet.cell(row=1, column=col)
            header = str(cell.value) if cell.value else f"Col{col}"
            headers.append(header)
            col_letter = chr(64 + col) if col <= 26 else f"Col{col}"
            print(f"  {col_letter}: {header}")
        
        # Show some sample rows to understand structure
        print(f"\nSample data rows:")
        for row in range(2, min(6, sheet.max_row + 1)):  # Show rows 2-5
            print(f"\nRow {row}:")
            for col in range(1, min(6, sheet.max_column + 1)):  # Show first 5 columns
                cell = sheet.cell(row=row, column=col)
                value = str(cell.value) if cell.value else ""
                col_letter = chr(64 + col) if col <= 26 else f"Col{col}"
                print(f"  {col_letter}: {value[:50]}")
        
        # Find which column contains package names
        print(f"\nSearching for package names...")
        package_column = None
        for col in range(1, min(sheet.max_column + 1, 10)):  # Check first 10 columns
            for row in range(2, min(10, sheet.max_row + 1)):  # Check first few data rows
                cell = sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    value = str(cell.value).lower()
                    if any(pkg in value for pkg in ['pyinstaller', 'pyjwt', 'requests', 'numpy', 'pandas']):
                        package_column = col
                        col_letter = chr(64 + col) if col <= 26 else f"Col{col}"
                        print(f"  Found package names in column {col_letter} (sample: {cell.value})")
                        break
            if package_column:
                break
        
        if not package_column:
            print("  No obvious package name column found")
            
            # Search specifically for pyinstaller and PyJWT across all columns and rows
            print(f"\nSearching for 'pyinstaller' and 'PyJWT' across all data...")
            found_locations = []
            for row in range(1, min(sheet.max_row + 1, 500)):  # Search first 500 rows
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        value = str(cell.value).lower()
                        if 'pyinstaller' in value or 'pyjwt' in value:
                            col_letter = chr(64 + col) if col <= 26 else f"Col{col}"
                            found_locations.append((row, col, col_letter, str(cell.value)))
            
            if found_locations:
                print(f"  Found target packages:")
                for row, col, col_letter, value in found_locations[:10]:  # Show first 10 matches
                    print(f"    Row {row}, Col {col_letter}: {value}")
            else:
                print(f"  Target packages not found in first 500 rows")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error examining file: {e}")

def main():
    """Main function"""
    manual_file = "/mnt/c/workspace/ihacpa-v2/Copy of 2025-07-23-updated v0.9.xlsx"
    automated_file = "/mnt/c/workspace/ihacpa-v2/test_fixes.xlsx"
    
    print("EXCEL STRUCTURE EXAMINATION")
    print("=" * 80)
    
    if Path(manual_file).exists():
        examine_structure(manual_file, "MANUAL FILE")
    else:
        print(f"❌ Manual file not found: {manual_file}")
    
    if Path(automated_file).exists():
        examine_structure(automated_file, "AUTOMATED FILE")
    else:
        print(f"❌ Automated file not found: {automated_file}")

if __name__ == "__main__":
    main()