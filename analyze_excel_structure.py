#!/usr/bin/env python3
"""
Analyze Excel structure and content to understand the IHACPA processing results
"""

import openpyxl
import sys
from pathlib import Path

def analyze_excel_structure(file_path):
    """Analyze the Excel file structure using openpyxl"""
    try:
        print(f"📊 Analyzing Excel Structure: {file_path}")
        print("=" * 80)
        
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        print(f"📋 Sheet name: {sheet.title}")
        print(f"📏 Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
        
        # Check first few rows to understand structure
        print(f"\n🔍 First 10 rows structure:")
        print("-" * 60)
        
        for row in range(1, min(11, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(24, sheet.max_column + 1)):  # Check first 23 columns (A-W)
                cell = sheet.cell(row=row, column=col)
                value = str(cell.value) if cell.value is not None else ""
                # Truncate long values
                if len(value) > 30:
                    value = value[:27] + "..."
                row_data.append(value)
            
            # Convert column numbers to letters for display
            col_letters = [chr(65 + i) for i in range(len(row_data))]
            print(f"Row {row:2d}: ", end="")
            for i, (letter, val) in enumerate(zip(col_letters, row_data)):
                if val.strip():
                    print(f"{letter}:{val[:15]} ", end="")
            print()
        
        # Look specifically at columns that should contain enhanced data
        enhanced_columns = [5, 6, 8, 11, 12, 13, 23]  # E, F, H, K, L, M, W (1-indexed)
        enhanced_letters = ['E', 'F', 'H', 'K', 'L', 'M', 'W']
        
        print(f"\n🔧 Enhanced Columns Analysis:")
        print("-" * 60)
        
        for col_num, col_letter in zip(enhanced_columns, enhanced_letters):
            if col_num <= sheet.max_column:
                # Count non-empty cells (excluding header rows)
                non_empty_count = 0
                sample_values = []
                
                for row in range(3, min(sheet.max_row + 1, 20)):  # Skip first 2 rows, check next 17
                    cell = sheet.cell(row=row, column=col_num)
                    if cell.value is not None and str(cell.value).strip():
                        non_empty_count += 1
                        if len(sample_values) < 3:
                            sample_values.append(str(cell.value)[:50])
                
                total_data_rows = max(0, sheet.max_row - 2)  # Exclude header rows
                fill_rate = (non_empty_count / total_data_rows * 100) if total_data_rows > 0 else 0
                
                print(f"Column {col_letter} (col {col_num}): {non_empty_count}/{total_data_rows} filled ({fill_rate:.1f}%)")
                if sample_values:
                    print(f"  Samples: {sample_values}")
                else:
                    print(f"  No data found")
            else:
                print(f"Column {col_letter}: Column doesn't exist")
        
        # Look for any recently modified cells (cells with formulas or special formatting)
        print(f"\n📈 Special Content Analysis:")
        print("-" * 60)
        
        formulas_found = 0
        hyperlinks_found = 0
        
        for row in range(1, min(sheet.max_row + 1, 50)):  # Check first 50 rows
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                
                # Check for formulas
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas_found += 1
                    if formulas_found <= 3:  # Show first 3 examples
                        col_letter = chr(64 + col)
                        print(f"  Formula in {col_letter}{row}: {cell.value[:50]}")
                
                # Check for hyperlinks
                if cell.hyperlink:
                    hyperlinks_found += 1
                    if hyperlinks_found <= 3:  # Show first 3 examples
                        col_letter = chr(64 + col)
                        print(f"  Hyperlink in {col_letter}{row}: {cell.hyperlink.target[:50]}")
        
        print(f"\n📊 Processing Summary:")
        print("-" * 60)
        print(f"✅ Total rows: {sheet.max_row}")
        print(f"✅ Total columns: {sheet.max_column}")
        print(f"🔧 Formulas found: {formulas_found}")
        print(f"🔗 Hyperlinks found: {hyperlinks_found}")
        
        # Determine if processing occurred
        if formulas_found > 0 or hyperlinks_found > 0:
            print("✅ Enhanced processing detected (formulas/hyperlinks found)!")
        else:
            # Check for any filled enhanced columns
            enhanced_data_found = False
            for col_num in enhanced_columns:
                if col_num <= sheet.max_column:
                    for row in range(3, min(sheet.max_row + 1, 10)):
                        cell = sheet.cell(row=row, column=col_num)
                        if cell.value is not None and str(cell.value).strip():
                            enhanced_data_found = True
                            break
                if enhanced_data_found:
                    break
            
            if enhanced_data_found:
                print("✅ Enhanced data found in target columns!")
            else:
                print("⚠️  No enhanced processing detected")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"❌ Error analyzing file: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main function"""
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        # Look for enhanced files
        enhanced_files = list(Path(".").glob("*enhanced*.xlsx"))
        if enhanced_files:
            file_path = str(enhanced_files[0])
            print(f"🎯 Auto-detected enhanced file: {file_path}")
        else:
            print("❌ No enhanced Excel file found. Please specify file path.")
            return
    
    if not Path(file_path).exists():
        print(f"❌ File not found: {file_path}")
        return
    
    analyze_excel_structure(file_path)

if __name__ == "__main__":
    main()