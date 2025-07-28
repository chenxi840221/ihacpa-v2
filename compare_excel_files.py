#!/usr/bin/env python3
"""
Compare manual reviewed results with automated results for IHACPA analysis
Focus on pyinstaller and PyJWT packages in rows 312 and 314
"""

import openpyxl
import sys
from pathlib import Path

def get_column_letter(col_num):
    """Convert column number to letter (1=A, 2=B, etc.)"""
    if col_num <= 26:
        return chr(64 + col_num)
    else:
        return chr(64 + ((col_num - 1) // 26)) + chr(65 + ((col_num - 1) % 26))

def analyze_cell_formatting(cell):
    """Analyze cell formatting including color"""
    formatting_info = {}
    
    # Check fill color
    if cell.fill and cell.fill.start_color:
        color = cell.fill.start_color
        if hasattr(color, 'rgb') and color.rgb and color.rgb != '00000000':
            formatting_info['fill_color'] = color.rgb
    
    # Check font color
    if cell.font and cell.font.color:
        color = cell.font.color
        if hasattr(color, 'rgb') and color.rgb and color.rgb != '00000000':
            formatting_info['font_color'] = color.rgb
    
    # Check if bold
    if cell.font and cell.font.bold:
        formatting_info['bold'] = True
    
    return formatting_info

def compare_excel_files(manual_file, automated_file):
    """Compare the two Excel files focusing on specific packages and columns"""
    print("IHACPA COMPARISON REPORT: Manual vs Automated Results")
    print("=" * 80)
    
    try:
        # Load both workbooks
        manual_wb = openpyxl.load_workbook(manual_file)
        automated_wb = openpyxl.load_workbook(automated_file)
        
        manual_sheet = manual_wb.active
        automated_sheet = automated_wb.active
        
        print(f"📊 Manual file: {manual_file}")
        print(f"   - Dimensions: {manual_sheet.max_row} rows x {manual_sheet.max_column} columns")
        print(f"📊 Automated file: {automated_file}")
        print(f"   - Dimensions: {automated_sheet.max_row} rows x {automated_sheet.max_column} columns")
        
        # Target columns: E, F, H, K, L, M, P, R, T, V, W
        target_columns = [5, 6, 8, 11, 12, 13, 16, 18, 20, 22, 23]
        column_letters = ['E', 'F', 'H', 'K', 'L', 'M', 'P', 'R', 'T', 'V', 'W']
        
        # First, let's identify the rows with pyinstaller and PyJWT
        print(f"\n🔍 SEARCHING FOR TARGET PACKAGES...")
        print("-" * 60)
        
        target_packages = ['pyinstaller', 'PyJWT']
        package_rows = {}
        
        # Search in column A (package names) in manual file
        for row in range(1, manual_sheet.max_row + 1):
            cell_a = manual_sheet.cell(row=row, column=1)  # Column A
            if cell_a.value:
                package_name = str(cell_a.value).strip().lower()
                for target_pkg in target_packages:
                    if target_pkg.lower() in package_name:
                        package_rows[target_pkg] = row
                        print(f"Found {target_pkg} in manual file at row {row}: {cell_a.value}")
        
        # Search in automated file as well
        automated_package_rows = {}
        for row in range(1, automated_sheet.max_row + 1):
            cell_a = automated_sheet.cell(row=row, column=1)  # Column A
            if cell_a.value:
                package_name = str(cell_a.value).strip().lower()
                for target_pkg in target_packages:
                    if target_pkg.lower() in package_name:
                        automated_package_rows[target_pkg] = row
                        print(f"Found {target_pkg} in automated file at row {row}: {cell_a.value}")
        
        if not package_rows:
            print("⚠️  Could not find target packages in manual file. Checking rows 312 and 314 anyway...")
            package_rows = {'package_312': 312, 'package_314': 314}
        
        if not automated_package_rows:
            print("⚠️  Could not find target packages in automated file. Using same rows...")
            automated_package_rows = package_rows.copy()
        
        print(f"\n📋 DETAILED COMPARISON BY PACKAGE AND COLUMN")
        print("=" * 80)
        
        # Compare each package
        for pkg_name in package_rows:
            manual_row = package_rows[pkg_name]
            automated_row = automated_package_rows.get(pkg_name, manual_row)
            
            print(f"\n🔍 PACKAGE: {pkg_name.upper()}")
            print(f"   Manual row: {manual_row}, Automated row: {automated_row}")
            print("-" * 60)
            
            # Get package name from both files
            manual_pkg_name = manual_sheet.cell(row=manual_row, column=1).value
            automated_pkg_name = automated_sheet.cell(row=automated_row, column=1).value
            
            print(f"Package Name:")
            print(f"  Manual:    {manual_pkg_name}")
            print(f"  Automated: {automated_pkg_name}")
            if str(manual_pkg_name) != str(automated_pkg_name):
                print(f"  ❌ DIFFERENCE: Package names don't match!")
            
            # Compare each target column
            for col_num, col_letter in zip(target_columns, column_letters):
                print(f"\n📊 Column {col_letter} (Position {col_num}):")
                
                # Get values from both files
                manual_cell = manual_sheet.cell(row=manual_row, column=col_num) if col_num <= manual_sheet.max_column else None
                automated_cell = automated_sheet.cell(row=automated_row, column=col_num) if col_num <= automated_sheet.max_column else None
                
                manual_value = manual_cell.value if manual_cell else "N/A"
                automated_value = automated_cell.value if automated_cell else "N/A"
                
                # Get formatting
                manual_format = analyze_cell_formatting(manual_cell) if manual_cell else {}
                automated_format = analyze_cell_formatting(automated_cell) if automated_cell else {}
                
                print(f"  Manual:    '{manual_value}'")
                if manual_format:
                    print(f"             Formatting: {manual_format}")
                print(f"  Automated: '{automated_value}'")
                if automated_format:
                    print(f"             Formatting: {automated_format}")
                
                # Compare values
                if str(manual_value) != str(automated_value):
                    print(f"  ❌ VALUE DIFFERENCE!")
                    
                # Compare formatting
                if manual_format != automated_format:
                    print(f"  ⚠️  FORMATTING DIFFERENCE!")
                
                if str(manual_value) == str(automated_value) and manual_format == automated_format:
                    print(f"  ✅ MATCH")
        
        # Get column headers for context
        print(f"\n📋 COLUMN HEADERS FOR REFERENCE")
        print("-" * 60)
        for col_num, col_letter in zip(target_columns, column_letters):
            manual_header = manual_sheet.cell(row=1, column=col_num).value if col_num <= manual_sheet.max_column else "N/A"
            automated_header = automated_sheet.cell(row=1, column=col_num).value if col_num <= automated_sheet.max_column else "N/A"
            print(f"Column {col_letter}: Manual='{manual_header}' | Automated='{automated_header}'")
        
        # Summary statistics
        print(f"\n📊 SUMMARY ANALYSIS")
        print("=" * 80)
        
        # Check if manual file has more data in target columns
        manual_data_count = 0
        automated_data_count = 0
        
        for pkg_name in package_rows:
            manual_row = package_rows[pkg_name]
            automated_row = automated_package_rows.get(pkg_name, manual_row)
            
            for col_num in target_columns:
                manual_cell = manual_sheet.cell(row=manual_row, column=col_num) if col_num <= manual_sheet.max_column else None
                automated_cell = automated_sheet.cell(row=automated_row, column=col_num) if col_num <= automated_sheet.max_column else None
                
                if manual_cell and manual_cell.value is not None and str(manual_cell.value).strip():
                    manual_data_count += 1
                if automated_cell and automated_cell.value is not None and str(automated_cell.value).strip():
                    automated_data_count += 1
        
        print(f"Data completeness:")
        print(f"  Manual file: {manual_data_count}/{len(package_rows) * len(target_columns)} cells populated")
        print(f"  Automated file: {automated_data_count}/{len(package_rows) * len(target_columns)} cells populated")
        
        if manual_data_count > automated_data_count:
            print(f"  ⚠️  Manual file has more data ({manual_data_count - automated_data_count} more cells)")
        elif automated_data_count > manual_data_count:
            print(f"  ✅ Automated file has more data ({automated_data_count - manual_data_count} more cells)")
        else:
            print(f"  ✅ Both files have equal data completeness")
        
        # Close workbooks
        manual_wb.close()
        automated_wb.close()
        
        return True
        
    except Exception as e:
        print(f"❌ Error comparing files: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main function"""
    manual_file = "/mnt/c/workspace/ihacpa-v2/Copy of 2025-07-23-updated v0.9.xlsx"
    automated_file = "/mnt/c/workspace/ihacpa-v2/test_fixes.xlsx"
    
    if not Path(manual_file).exists():
        print(f"❌ Manual file not found: {manual_file}")
        return
    
    if not Path(automated_file).exists():
        print(f"❌ Automated file not found: {automated_file}")
        return
    
    compare_excel_files(manual_file, automated_file)

if __name__ == "__main__":
    main()