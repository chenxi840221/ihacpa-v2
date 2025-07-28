"""
Enhanced Excel Handler for IHACPA v2.0

Provides comprehensive Excel file operations with advanced features including
color highlighting, change tracking, validation, and multiple format export.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any, Union
import logging
import json
import csv
import shutil
from ..config import Config


class ExcelHandler:
    """Enhanced Excel file handler with advanced formatting and validation"""
    
    def __init__(self, file_path: Union[str, Path], config: Config):
        """
        Initialize Excel handler.
        
        Args:
            file_path: Path to Excel file
            config: Application configuration
        """
        self.file_path = Path(file_path)
        self.config = config
        self.workbook: Optional[openpyxl.Workbook] = None
        self.worksheet: Optional[openpyxl.worksheet.worksheet.Worksheet] = None
        self.logger = logging.getLogger(__name__)
        
        # Get column mapping from config
        self.column_mapping = config.excel.column_mapping
        
        # Color definitions for highlighting changes (matching old version)
        self.colors = {
            'updated': PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),  # Light blue
            'new_data': PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),  # Light green
            'security_risk': PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),  # Light red
            'version_update': PatternFill(start_color="FFF0E6", end_color="FFF0E6", fill_type="solid"),  # Light orange
            'github_added': PatternFill(start_color="F0E6FF", end_color="F0E6FF", fill_type="solid"),  # Light purple
            'not_available': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red for Not Available
        }
        
        # Font color definitions that complement the fill colors (matching old version)
        self.fonts = {
            'updated': Font(color="0066CC", bold=True),           # Bright blue (bold) for light blue background
            'new_data': Font(color="006600", bold=True),          # Medium green (bold) for light green background - better contrast
            'security_risk': Font(color="CC0000", bold=True),     # Bright red (bold) for light red background
            'version_update': Font(color="FF6600", bold=True),    # Bright orange (bold) for light orange background
            'github_added': Font(color="6600CC", bold=True),      # Bright purple (bold) for light purple background
            'not_available': Font(color="FFFFFF", bold=True),     # White (bold) for red background
            'default': Font(color="000000", bold=False),          # Black for white/no background
        }
        
        # Track changes for reporting
        self.changes_made: List[Dict[str, Any]] = []
        
    def load_workbook(self) -> bool:
        """
        Load Excel workbook and get active worksheet.
        
        Returns:
            True if loaded successfully, False otherwise
        """
        try:
            if not self.file_path.exists():
                self.logger.error(f"Excel file not found: {self.file_path}")
                return False
                
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            self.logger.info(f"Successfully loaded Excel file: {self.file_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading Excel file: {e}")
            return False
    
    def validate_file_structure(self) -> Tuple[bool, List[str]]:
        """
        Validate Excel file structure.
        
        Returns:
            Tuple of (is_valid, list_of_errors)
        """
        errors = []
        
        if not self.worksheet:
            errors.append("Worksheet not loaded")
            return False, errors
        
        # Check if file has minimum required columns
        if self.worksheet.max_column < len(self.column_mapping):
            errors.append(f"Expected at least {len(self.column_mapping)} columns, found {self.worksheet.max_column}")
        
        # Check if file has data rows
        if self.worksheet.max_row < self.config.excel.data_start_row:
            errors.append(f"No data rows found (expected data starting from row {self.config.excel.data_start_row})")
        
        # Validate header row exists
        header_row = self.config.excel.header_row
        if header_row > self.worksheet.max_row:
            errors.append(f"Header row {header_row} not found in file")
        
        return len(errors) == 0, errors
    
    def create_backup(self) -> Optional[Path]:
        """
        Create backup of original file.
        
        Returns:
            Path to backup file if successful, None otherwise
        """
        if not self.config.excel.backup_original:
            return None
            
        try:
            backup_dir = Path(self.config.output.backup_directory)
            backup_dir.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"{self.file_path.stem}_backup_{timestamp}{self.file_path.suffix}"
            backup_path = backup_dir / backup_filename
            
            shutil.copy2(self.file_path, backup_path)
            self.logger.info(f"Backup created: {backup_path}")
            return backup_path
            
        except Exception as e:
            self.logger.error(f"Failed to create backup: {e}")
            return None
    
    def get_package_count(self) -> int:
        """
        Get total number of packages in the Excel file.
        
        Returns:
            Number of packages found
        """
        if not self.worksheet:
            return 0
            
        count = 0
        package_col = self.column_mapping.get('package_name', 2)
        
        for row in range(self.config.excel.data_start_row, self.worksheet.max_row + 1):
            package_name = self.worksheet.cell(row=row, column=package_col).value
            if package_name and str(package_name).strip():
                count += 1
        return count
    
    def get_package_data(self, row_number: int) -> Dict[str, Any]:
        """
        Get package data for a specific row.
        
        Args:
            row_number: Row number to read
            
        Returns:
            Dictionary with package data
        """
        if not self.worksheet:
            return {}
            
        package_data = {}
        for field, column in self.column_mapping.items():
            cell_value = self.worksheet.cell(row=row_number, column=column).value
            # Clean up the value
            if isinstance(cell_value, str):
                cell_value = cell_value.strip()
            package_data[field] = cell_value
            
        return package_data
    
    def get_all_packages(self) -> List[Dict[str, Any]]:
        """
        Get all package data from Excel file.
        
        Returns:
            List of dictionaries with package data
        """
        if not self.worksheet:
            return []
            
        packages = []
        package_col = self.column_mapping.get('package_name', 2)
        
        for row in range(self.config.excel.data_start_row, self.worksheet.max_row + 1):
            package_name = self.worksheet.cell(row=row, column=package_col).value
            if package_name and str(package_name).strip():
                package_data = self.get_package_data(row)
                package_data['row_number'] = row
                packages.append(package_data)
                
        return packages
    
    def find_package_by_name(self, package_name: str) -> Optional[Dict[str, Any]]:
        """
        Find package by name in Excel file.
        
        Args:
            package_name: Name of package to find
            
        Returns:
            Package data dictionary if found, None otherwise
        """
        packages = self.get_all_packages()
        for package in packages:
            if package.get('package_name', '').lower() == package_name.lower():
                return package
        return None
    
    def update_package_data(self, row_number: int, updates: Dict[str, Any], 
                          highlight_changes: bool = True) -> bool:
        """
        Update package data for a specific row with optional highlighting.
        
        Args:
            row_number: Row number to update
            updates: Dictionary of field updates
            highlight_changes: Whether to apply color highlighting
            
        Returns:
            True if updated successfully, False otherwise
        """
        if not self.worksheet:
            return False
            
        try:
            for field, value in updates.items():
                if field in self.column_mapping:
                    column = self.column_mapping[field]
                    cell = self.worksheet.cell(row=row_number, column=column)
                    
                    # Store original value for change tracking
                    original_value = cell.value
                    
                    # Handle datetime objects
                    if isinstance(value, datetime):
                        # Remove timezone info and microseconds for Excel
                        value = value.replace(tzinfo=None, microsecond=0)
                    
                    # Only update if value has changed
                    if original_value != value:
                        cell.value = value
                        
                        # Track the change
                        self.changes_made.append({
                            'row': row_number,
                            'field': field,
                            'old_value': original_value,
                            'new_value': value,
                            'timestamp': datetime.now()
                        })
                        
                        # Apply highlighting if requested
                        if highlight_changes:
                            self._apply_cell_formatting(cell, field, value, original_value)
                        
                        self.logger.debug(f"Updated {field} in row {row_number}: {original_value} → {value}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error updating package data: {e}")
            return False
    
    def _apply_cell_formatting(self, cell, field: str, new_value: Any, old_value: Any):
        """Apply formatting to a cell based on the type of change"""
        color_type = self._determine_color_type(field, new_value, old_value)
        
        if color_type:
            # Apply fill color
            cell.fill = self.colors[color_type]
            
            # Apply font
            font_style = self.fonts.get(color_type, self.fonts['default'])
            existing_font = cell.font
            cell.font = Font(
                color=font_style.color,
                bold=font_style.bold,
                italic=existing_font.italic,
                size=existing_font.size or 11,
                name=existing_font.name or 'Calibri'
            )
            
            # Ensure proper alignment
            cell.alignment = Alignment(
                wrap_text=True,
                horizontal=cell.alignment.horizontal or 'center',
                vertical=cell.alignment.vertical or 'center'
            )
    
    def _determine_color_type(self, field: str, new_value: Any, old_value: Any) -> Optional[str]:
        """Determine appropriate color type based on field and values"""
        if old_value is None and new_value is not None:
            return 'new_data'
        
        # Security-related fields
        field_str = str(field).lower()
        if 'result' in field_str and new_value:
            value_str = str(new_value).upper()
            if any(term in value_str for term in ['CRITICAL', 'VULNERABLE', 'HIGH']):
                return 'security_risk'
            elif 'AI' in value_str or 'ENHANCED' in value_str:
                return 'ai_enhanced'
        
        # Version updates
        if 'version' in field_str and old_value != new_value:
            return 'version_update'
        
        # Default update highlighting
        return 'updated'
    
    def add_summary_sheet(self, summary_data: Dict[str, Any]) -> bool:
        """
        Add a summary sheet with scan results.
        
        Args:
            summary_data: Dictionary containing summary information
            
        Returns:
            True if added successfully, False otherwise
        """
        try:
            # Create or get summary sheet
            if "Summary" in self.workbook.sheetnames:
                summary_sheet = self.workbook["Summary"]
                summary_sheet.delete_rows(1, summary_sheet.max_row)
            else:
                summary_sheet = self.workbook.create_sheet("Summary", 0)
            
            # Add header
            summary_sheet.cell(row=1, column=1, value="IHACPA v2.0 Scan Summary")
            summary_sheet.cell(row=1, column=1).font = Font(size=16, bold=True)
            
            row = 3
            for key, value in summary_data.items():
                summary_sheet.cell(row=row, column=1, value=key)
                summary_sheet.cell(row=row, column=2, value=str(value))
                row += 1
            
            # Adjust column widths
            summary_sheet.column_dimensions['A'].width = 30
            summary_sheet.column_dimensions['B'].width = 20
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error adding summary sheet: {e}")
            return False
    
    def save_workbook(self, output_path: Optional[Path] = None) -> bool:
        """
        Save the workbook.
        
        Args:
            output_path: Optional custom output path
            
        Returns:
            True if saved successfully, False otherwise
        """
        try:
            save_path = output_path or self.file_path
            
            # Create output directory if needed
            save_path.parent.mkdir(parents=True, exist_ok=True)
            
            self.workbook.save(save_path)
            self.logger.info(f"Excel file saved: {save_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error saving Excel file: {e}")
            return False
    
    def update_cell(self, row_number: int, column_field: str, value: Any, 
                   color: Optional[str] = None, font: Optional[str] = None) -> bool:
        """
        Update a single cell with value and formatting.
        
        Args:
            row_number: Row number to update
            column_field: Column field name from mapping
            value: Value to set
            color: Optional color key for cell background
            font: Optional font key for text formatting
            
        Returns:
            True if updated successfully, False otherwise
        """
        if not self.worksheet or column_field not in self.column_mapping:
            return False
            
        try:
            column = self.column_mapping[column_field]
            cell = self.worksheet.cell(row=row_number, column=column)
            
            # Store original value for change tracking
            original_value = cell.value
            
            # Set new value
            cell.value = value
            
            # Apply formatting if specified
            if color and color in self.colors:
                cell.fill = self.colors[color]
            
            if font and font in self.fonts:
                cell.font = self.fonts[font]
            
            # Track change
            self.changes_made.append({
                'row': row_number,
                'column': column_field,
                'old_value': original_value,
                'new_value': value,
                'timestamp': datetime.now().isoformat()
            })
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error updating cell {column_field} at row {row_number}: {e}")
            return False
    
    def add_hyperlink(self, row_number: int, column_field: str, url: str, display_text: Optional[str] = None) -> bool:
        """
        Add hyperlink to a cell.
        
        Args:
            row_number: Row number
            column_field: Column field name from mapping
            url: URL for the hyperlink
            display_text: Optional display text (uses URL if not provided)
            
        Returns:
            True if added successfully, False otherwise
        """
        if not self.worksheet or column_field not in self.column_mapping:
            return False
            
        try:
            from openpyxl.worksheet.hyperlink import Hyperlink
            
            column = self.column_mapping[column_field]
            cell = self.worksheet.cell(row=row_number, column=column)
            
            # Set display text
            cell.value = display_text or url
            
            # Add hyperlink
            cell.hyperlink = url
            
            # Apply hyperlink styling
            cell.font = Font(color="0066CC", underline="single")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error adding hyperlink to {column_field} at row {row_number}: {e}")
            return False

    def export_to_csv(self, output_path: Path) -> bool:
        """
        Export data to CSV format.
        
        Args:
            output_path: Path for CSV output
            
        Returns:
            True if exported successfully, False otherwise
        """
        try:
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                header = list(self.column_mapping.keys())
                writer.writerow(header)
                
                # Write data
                for row in range(self.config.excel.data_start_row, self.worksheet.max_row + 1):
                    row_data = []
                    for field in header:
                        column = self.column_mapping[field]
                        value = self.worksheet.cell(row=row, column=column).value
                        row_data.append(str(value) if value is not None else '')
                    writer.writerow(row_data)
            
            self.logger.info(f"CSV export completed: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error exporting to CSV: {e}")
            return False
    
    def export_to_json(self, output_path: Path) -> bool:
        """
        Export data to JSON format.
        
        Args:
            output_path: Path for JSON output
            
        Returns:
            True if exported successfully, False otherwise
        """
        try:
            packages = self.get_all_packages()
            
            # Remove row_number from export data
            for package in packages:
                package.pop('row_number', None)
            
            export_data = {
                'metadata': {
                    'export_date': datetime.now().isoformat(),
                    'source_file': str(self.file_path),
                    'total_packages': len(packages),
                    'ihacpa_version': self.config.app.version
                },
                'packages': packages,
                'changes_made': self.changes_made
            }
            
            with open(output_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(export_data, jsonfile, indent=2, default=str)
            
            self.logger.info(f"JSON export completed: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error exporting to JSON: {e}")
            return False
    
    def get_changes_summary(self) -> Dict[str, Any]:
        """
        Get summary of changes made.
        
        Returns:
            Dictionary with change statistics
        """
        if not self.changes_made:
            return {'total_changes': 0}
        
        changes_by_field = {}
        changes_by_row = {}
        
        for change in self.changes_made:
            field = change['field']
            row = change['row']
            
            changes_by_field[field] = changes_by_field.get(field, 0) + 1
            changes_by_row[row] = changes_by_row.get(row, 0) + 1
        
        return {
            'total_changes': len(self.changes_made),
            'changes_by_field': changes_by_field,
            'rows_modified': len(changes_by_row),
            'most_changed_fields': sorted(changes_by_field.items(), key=lambda x: x[1], reverse=True)[:5]
        }
    
    def close(self):
        """Clean up resources"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None