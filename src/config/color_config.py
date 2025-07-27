"""
Color Configuration for IHACPA v2.0

Defines consistent color coding for Excel cells based on manual review standards.
"""

from typing import Dict, Tuple
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.colors import Color


class ExcelColors:
    """Excel color definitions matching manual review standards"""
    
    # Color definitions (ARGB format without FF prefix for openpyxl)
    COLORS = {
        # Status colors
        'safe': 'E6FFE6',           # Light green - No issues found
        'safe_alt': 'C6EFCE',       # Alternative light green
        'vulnerable': 'FFE6E6',     # Light red/pink - Vulnerability detected
        'manual_review': 'E6F3FF',  # Light blue/pink - Manual review required
        'warning': 'FFF2CC',        # Light yellow - Warning/caution
        'info': 'E6E6FA',          # Light lavender - Informational
        'error': 'FFCCCC',         # Darker red - Error/critical
        'neutral': 'F2F2F2',       # Light gray - Neutral/no data
        'new_data': 'FFE6CC',      # Light orange - New/updated data
        
        # Special case colors
        'maintenance': 'FFE6F5',    # Light pink - Maintenance mode
        'deprecated': 'FFE6E6',     # Light red - Deprecated package
        'false_positive': 'E6FFF0', # Very light green - False positive filtered
    }
    
    # Font colors
    FONT_COLORS = {
        'default': '000000',        # Black
        'safe': '006100',           # Dark green
        'vulnerable': '9C0006',     # Dark red
        'manual_review': '1F4788',  # Dark blue
        'warning': '9C6500',        # Dark orange
        'info': '5F5F5F',          # Dark gray
        'hyperlink': '0563C1',      # Hyperlink blue
    }
    
    @classmethod
    def get_fill(cls, color_type: str) -> PatternFill:
        """Get PatternFill object for color type"""
        color_code = cls.COLORS.get(color_type, cls.COLORS['neutral'])
        return PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')
    
    @classmethod
    def get_font(cls, color_type: str, bold: bool = False, underline: bool = False) -> Font:
        """Get Font object for color type"""
        font_color = cls.FONT_COLORS.get(color_type, cls.FONT_COLORS['default'])
        return Font(color=font_color, bold=bold, underline='single' if underline else None)
    
    @classmethod
    def get_color_for_vulnerability_status(cls, status: Dict[str, any]) -> str:
        """
        Determine color based on vulnerability status.
        
        Args:
            status: Dictionary with vulnerability information
            
        Returns:
            Color type string
        """
        # Check various status indicators
        if status.get('requires_manual_review'):
            return 'manual_review'
        
        if status.get('vulnerabilities_found', 0) > 0:
            if status.get('current_version_affected'):
                return 'vulnerable'
            else:
                return 'safe'  # Vulnerabilities exist but not in current version
        
        if status.get('false_positives_filtered', 0) > 0:
            return 'false_positive'
        
        if status.get('is_deprecated'):
            return 'deprecated'
        
        if status.get('maintenance_mode'):
            return 'maintenance'
        
        if status.get('no_data'):
            return 'neutral'
        
        return 'safe'  # Default to safe if no issues
    
    @classmethod
    def get_color_mapping_for_column(cls, column: str) -> Dict[str, str]:
        """
        Get color mapping rules for specific columns.
        
        Args:
            column: Column letter
            
        Returns:
            Dictionary mapping status to color
        """
        # Column-specific color mappings
        mappings = {
            'P': {  # NIST NVD
                'no_vulnerabilities': 'safe',
                'vulnerabilities_not_affecting_current': 'safe',
                'vulnerabilities_affecting_current': 'vulnerable',
                'manual_review_required': 'manual_review',
                'api_error': 'neutral',
                'not_found': 'safe_alt'
            },
            'R': {  # MITRE
                'no_vulnerabilities': 'safe',
                'vulnerabilities_not_affecting_current': 'safe',
                'vulnerabilities_affecting_current': 'vulnerable',
                'manual_review_required': 'manual_review',
                'api_error': 'neutral',
                'not_found': 'safe_alt'
            },
            'T': {  # SNYK
                'no_vulnerabilities': 'safe_alt',
                'vulnerabilities_not_affecting_current': 'safe',
                'vulnerabilities_affecting_current': 'vulnerable',
                'not_in_database': 'safe_alt',
                'api_error': 'neutral'
            },
            'V': {  # ExploitDB
                'no_exploits': 'safe_alt',
                'exploits_found': 'vulnerable',
                'not_found': 'safe_alt',
                'api_error': 'neutral'
            },
            'W': {  # Recommendations
                'security_risk': 'vulnerable',
                'manual_review': 'manual_review',
                'safe_to_update': 'safe',
                'version_updates': 'warning',
                'additional_info': 'info'
            }
        }
        
        return mappings.get(column, {})


class ColorCodingRules:
    """Rules for applying consistent color coding"""
    
    @staticmethod
    def apply_cell_formatting(cell, status: Dict[str, any], column: str):
        """
        Apply formatting to a cell based on status and column.
        
        Args:
            cell: openpyxl cell object
            status: Status information dictionary
            column: Column letter
        """
        # Determine color type
        color_type = ExcelColors.get_color_for_vulnerability_status(status)
        
        # Apply fill
        cell.fill = ExcelColors.get_fill(color_type)
        
        # Apply font
        is_critical = status.get('severity') in ['CRITICAL', 'HIGH']
        cell.font = ExcelColors.get_font(
            color_type, 
            bold=is_critical,
            underline=status.get('has_link', False)
        )
        
        # Apply alignment
        cell.alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=True
        )