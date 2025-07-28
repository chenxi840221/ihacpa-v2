#!/usr/bin/env python3
"""
Find the exact rows for pyinstaller and PyJWT packages in both Excel files
"""

import openpyxl
from pathlib import Path

def find_packages(file_path, file_name):
    """Find target packages in the specified Excel file"""
    print(f"\n🔍 SEARCHING IN {file_name}")
    print("-" * 60)
    
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        target_packages = ['pyinstaller', 'pyjwt']
        found_packages = {}
        
        # Search through all rows
        for row in range(1, sheet.max_row + 1):
            cell_a = sheet.cell(row=row, column=1)  # Column A (package names)
            if cell_a.value:
                package_name = str(cell_a.value).strip().lower()
                
                for target_pkg in target_packages:
                    if target_pkg in package_name:
                        found_packages[target_pkg] = {
                            'row': row,
                            'name': str(cell_a.value),
                            'exact_match': package_name == target_pkg
                        }
                        print(f"Row {row:3d}: {cell_a.value} {'(EXACT)' if package_name == target_pkg else '(PARTIAL)'}")
        
        # If no exact matches, show rows around 312 and 314 for context
        if not found_packages:
            print("No target packages found. Showing context around rows 312-314:")
            for row in range(310, 316):
                if row <= sheet.max_row:
                    cell_a = sheet.cell(row=row, column=1)
                    print(f"Row {row:3d}: {cell_a.value}")
        
        wb.close()
        return found_packages
        
    except Exception as e:
        print(f"❌ Error searching file: {e}")
        return {}

def main():
    """Main function"""
    manual_file = "/mnt/c/workspace/ihacpa-v2/Copy of 2025-07-23-updated v0.9.xlsx"
    automated_file = "/mnt/c/workspace/ihacpa-v2/test_fixes.xlsx"
    
    print("PACKAGE SEARCH REPORT")
    print("=" * 80)
    
    if Path(manual_file).exists():
        manual_results = find_packages(manual_file, "MANUAL FILE")
    else:
        print(f"❌ Manual file not found: {manual_file}")
        manual_results = {}
    
    if Path(automated_file).exists():
        automated_results = find_packages(automated_file, "AUTOMATED FILE")
    else:
        print(f"❌ Automated file not found: {automated_file}")
        automated_results = {}
    
    print(f"\n📊 SUMMARY")
    print("=" * 60)
    for pkg in ['pyinstaller', 'pyjwt']:
        print(f"\n{pkg.upper()}:")
        if pkg in manual_results:
            print(f"  Manual:    Row {manual_results[pkg]['row']} - {manual_results[pkg]['name']}")
        else:
            print(f"  Manual:    Not found")
        
        if pkg in automated_results:
            print(f"  Automated: Row {automated_results[pkg]['row']} - {automated_results[pkg]['name']}")
        else:
            print(f"  Automated: Not found")

if __name__ == "__main__":
    main()